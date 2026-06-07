import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from pathlib import Path

HISTORIAL_PATH = "reportes/historial_arca.xlsx"


def agregar_al_historial(datos: dict):
    Path("reportes").mkdir(exist_ok=True)

    if Path(HISTORIAL_PATH).exists():
        wb = openpyxl.load_workbook(HISTORIAL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial Pedidos"
        rojo = PatternFill("solid", fgColor="C00000")
        font_blanco = Font(bold=True, color="FFFFFF", size=11)
        centro = Alignment(horizontal="center")
        headers = ["Fecha", "Comercio", "Cliente", "CP", "Productos", "Subtotal", "Impuestos", "Total", "Envio"]
        ws.append(headers)
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col)
            cell.font = font_blanco
            cell.fill = rojo
            cell.alignment = centro
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 50
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 28

    fecha = datetime.fromisoformat(datos["fecha"]).strftime("%d/%m/%Y %H:%M")
    comercio = datos.get("objetivo", "Desconocido")[:40]
    productos_texto = ", ".join([
        f"{p['nombre']} (${p['precio']:.2f})"
        for p in datos["productos"]
    ])

    ws.append([
        fecha,
        comercio,
        datos["cliente"],
        datos["zip"],
        productos_texto,
        datos["subtotal"],
        datos["impuestos"],
        datos["total"],
        datos["envio"],
    ])

    fila = ws.max_row
    for col in [6, 7, 8]:
        ws.cell(row=fila, column=col).number_format = '"$"#,##0.00'

    wb.save(HISTORIAL_PATH)
    print(f"  📊 Historial actualizado: {HISTORIAL_PATH}")
    return HISTORIAL_PATH