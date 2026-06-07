import json
import re
import smtplib
import os
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv()


def extraer_datos_pedido(reporte_path: str = "sesiones/reporte.json") -> dict:
    with open(reporte_path, encoding="utf-8") as f:
        reporte = json.load(f)

    resultado = reporte.get("resultado", "")

    # Extraer productos SIN duplicados
    productos = []
    nombres_vistos = set()
    patron = r"\d+\.\s+(.+?)\s+-\s+\$([0-9.]+)"
    for match in re.finditer(patron, resultado):
        nombre = match.group(1).strip()
        precio = float(match.group(2))
        if nombre not in nombres_vistos:
            nombres_vistos.add(nombre)
            productos.append({"nombre": nombre, "precio": precio})

    # Extraer totales con fallback a 0
    def extraer_monto(patron_texto):
        m = re.search(patron_texto, resultado)
        return float(m.group(1)) if m else 0.0

    subtotal  = extraer_monto(r"Subtotal.*?\$([0-9.]+)")
    impuestos = extraer_monto(r"[Ii]mpuestos.*?\$([0-9.]+)")
    total     = extraer_monto(r"TOTAL.*?\$([0-9.]+)")

    # Si no hay total, calcularlo
    if total == 0 and productos:
        total = sum(p["precio"] for p in productos)

    # Extraer info de envío
    cliente_m = re.search(r"Dirección:\s+(.+?),", resultado)
    zip_m     = re.search(r"Código Postal\s+(\d+)", resultado)
    envio_m   = re.search(r"Método:\s+(.+?)(?:\\n|\n|$)", resultado)

    cliente = cliente_m.group(1).strip() if cliente_m else "Cliente"
    zip_code = zip_m.group(1) if zip_m else ""
    envio = envio_m.group(1).strip() if envio_m else "Envío estándar"

    return {
        "productos": productos,
        "subtotal": subtotal,
        "impuestos": impuestos,
        "total": total,
        "cliente": cliente,
        "zip": zip_code,
        "envio": envio,
        "fecha": reporte.get("fecha", datetime.now().isoformat()),
        "objetivo": reporte.get("objetivo", "Pedido"),
    }


def generar_excel(datos: dict, output_path: str = "reportes/pedido_arca.xlsx") -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedido Arca Continental"

    rojo      = PatternFill("solid", fgColor="C00000")
    gris      = PatternFill("solid", fgColor="F2F2F2")
    gris_claro= PatternFill("solid", fgColor="FAFAFA")
    font_blanco = Font(bold=True, color="FFFFFF", size=13)
    font_bold   = Font(bold=True)
    centro      = Alignment(horizontal="center", vertical="center")
    derecha     = Alignment(horizontal="right")

    # Título
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:E1")
    ws["A1"] = "ARCA CONTINENTAL — REGISTRO DE PEDIDO"
    ws["A1"].font = font_blanco
    ws["A1"].fill = rojo
    ws["A1"].alignment = centro

    # Info del pedido
    fecha = datetime.fromisoformat(datos["fecha"]).strftime("%d/%m/%Y %H:%M:%S")
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Fecha: {fecha}  |  Cliente: {datos['cliente']}  |  CP: {datos['zip']}  |  Envío: {datos['envio']}"
    ws["A2"].fill = gris
    ws["A2"].alignment = centro

    ws.append([])  # fila vacía

    # Headers
    headers = ["#", "Producto", "Precio Unitario", "Cantidad", "Total"]
    ws.append(headers)
    for col in range(1, 6):
        cell = ws.cell(row=4, column=col)
        cell.font = font_bold
        cell.fill = gris
        cell.alignment = centro

    # Productos
    for i, prod in enumerate(datos["productos"], 1):
        ws.append([
            i,
            prod["nombre"],
            prod["precio"],
            1,
            prod["precio"]
        ])
        ws.cell(row=4+i, column=3).number_format = '"$"#,##0.00'
        ws.cell(row=4+i, column=5).number_format = '"$"#,##0.00'
        if i % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=4+i, column=col).fill = gris_claro

    # Línea separadora
    ws.append([])

    # Totales
    fila_base = 5 + len(datos["productos"])
    totales = [
        ("Subtotal:", datos["subtotal"]),
        ("Impuestos:", datos["impuestos"]),
        ("TOTAL:", datos["total"]),
    ]
    for label, valor in totales:
        fila_base += 1
        ws.cell(row=fila_base, column=4, value=label).font = font_bold
        cell_valor = ws.cell(row=fila_base, column=5, value=valor)
        cell_valor.number_format = '"$"#,##0.00'
        cell_valor.font = font_bold if label == "TOTAL:" else None
        cell_valor.alignment = derecha

    # Anchos
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16

    Path("reportes").mkdir(exist_ok=True)
    wb.save(output_path)
    print(f"  📊 Excel: {output_path}")
    return output_path


def generar_ticket_html(datos: dict) -> str:
    fecha = datetime.fromisoformat(datos["fecha"]).strftime("%d/%m/%Y %H:%M")
    filas = "".join([
        f"<tr><td style='padding:8px;border-bottom:1px solid #eee'>{p['nombre']}</td>"
        f"<td style='padding:8px;border-bottom:1px solid #eee;text-align:right'>${p['precio']:.2f}</td></tr>"
        for p in datos["productos"]
    ])
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Ticket de compra</title></head>
<body style="margin:0;background:#f5f5f5;font-family:Arial,sans-serif">
<div style="max-width:580px;margin:30px auto;background:white;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.1)">
  <div style="background:#C00000;padding:24px;text-align:center">
    <h1 style="color:white;margin:0;font-size:22px">🏪 Arca Continental</h1>
    <p style="color:#ffcccc;margin:6px 0 0">Confirmación de pedido</p>
  </div>
  <div style="padding:28px">
    <p style="font-size:15px">Hola <b>{datos['cliente']}</b>,</p>
    <p>Tu pedido fue procesado el <b>{fecha}</b>. Aquí está tu resumen:</p>
    <table style="width:100%;border-collapse:collapse;margin:16px 0">
      <tr style="background:#f2f2f2">
        <th style="padding:8px;text-align:left">Producto</th>
        <th style="padding:8px;text-align:right">Precio</th>
      </tr>
      {filas}
    </table>
    <table style="width:100%;margin-top:8px">
      <tr><td style="padding:4px">Subtotal</td><td style="text-align:right">${datos['subtotal']:.2f}</td></tr>
      <tr><td style="padding:4px">Impuestos</td><td style="text-align:right">${datos['impuestos']:.2f}</td></tr>
      <tr style="font-size:17px;font-weight:bold;color:#C00000">
        <td style="padding:8px 4px">TOTAL</td>
        <td style="text-align:right">${datos['total']:.2f}</td>
      </tr>
    </table>
    <div style="margin-top:20px;padding:14px;background:#fff8f8;border-left:4px solid #C00000;border-radius:4px">
      <p style="margin:0"><b>📦 Envío:</b> {datos['envio']}</p>
      <p style="margin:6px 0 0"><b>📍 CP:</b> {datos['zip']}</p>
    </div>
    <p style="margin-top:20px;color:#555">Tu pedido será surtido por Arca Continental. ¡Gracias por tu compra!</p>
  </div>
  <div style="background:#f2f2f2;padding:12px;text-align:center;font-size:12px;color:#888">
    Hack4Her 2026 · Always on Shelf · Arca Continental
  </div>
</div>
</body></html>"""


def enviar_ticket(datos: dict, email_cliente: str, excel_path: str):
    remitente = os.getenv("EMAIL_REMITENTE", "")
    password  = os.getenv("EMAIL_PASSWORD", "")
    fecha     = datetime.fromisoformat(datos["fecha"]).strftime("%d/%m/%Y %H:%M")
    html      = generar_ticket_html(datos)

    # Siempre guardar el HTML localmente
    ticket_path = "reportes/ticket.html"
    with open(ticket_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  🎫 Ticket HTML: {ticket_path}")

    if not remitente or not password:
        print(f"  ⚠️  Agrega EMAIL_REMITENTE y EMAIL_PASSWORD al .env para enviar por correo")
        return

    msg = MIMEMultipart("mixed")
    msg["From"]    = remitente
    msg["To"]      = email_cliente
    msg["Subject"] = f"✅ Tu pedido Arca Continental — {fecha} — ${datos['total']:.2f}"
    msg.attach(MIMEText(html, "html"))

    with open(excel_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment; filename=pedido_arca.xlsx")
        msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(remitente, password)
            server.sendmail(remitente, email_cliente, msg.as_string())
        print(f"  📧 Ticket enviado a {email_cliente}")
    except Exception as e:
        print(f"  ⚠️  Email no enviado: {e}")


def generar_todo(email_cliente: str, reporte_path: str = "sesiones/reporte.json"):
    print("\n📊 Generando reporte...")
    datos      = extraer_datos_pedido(reporte_path)
    excel_path = generar_excel(datos)
    import sys, os; sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath("reportes/generar_reporte.py")))); from reportes.historial import agregar_al_historial
    agregar_al_historial(datos)
    from reportes.google_sheets import agregar_a_sheets
    agregar_a_sheets(datos)
    from reportes.database import agregar_a_db
    agregar_a_db(datos)
    enviar_ticket(datos, email_cliente, excel_path)
    print(f"\n✅ Listo")
    print(f"   Cliente  : {datos['cliente']}")
    print(f"   Productos: {len(datos['productos'])}")
    print(f"   Total    : ${datos['total']:.2f}")
    return datos


if __name__ == "__main__":
    import sys
    email = sys.argv[1] if len(sys.argv) > 1 else input("Email del cliente: ").strip()
    generar_todo(email)
    