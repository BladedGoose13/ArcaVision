"""
postprocessing/reporte.py
--------------------------
Genera tres artefactos al terminar la ejecución del agente:

  1. Excel de compras   — productos, cantidad, precio, datos Arca
  2. PDF de reporte IA  — qué hizo, errores, iteraciones, info para ingenieros
  3. Excel de errores   — tipo de error + frecuencia (para gráficas)

Además mantiene el ticket HTML y el historial acumulativo existentes.
"""

import json
import re
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, KeepTogether,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


# ── Paleta Arca Continental ───────────────────────────────────────────────────
ROJO      = "C8102E"
ROJO_RL   = colors.HexColor("#C8102E")
GRIS      = "F2F2F2"
GRIS_CLARO = "FAFAFA"
OK_COLOR  = "E6F4EC"
ERR_COLOR = "FDECEA"
WARN_COLOR = "FEF3E2"
AZUL_CLARO = "EFF6FF"

Path("reportes").mkdir(exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════════════
# 1. Excel de compras
# ═══════════════════════════════════════════════════════════════════════════════

def generar_excel_compras(datos: dict,
                          output_path: str = "reportes/compras_arcavision.xlsx") -> str:
    """
    Hoja 1 — Detalle de productos comprados con precio, cantidad y datos Arca.
    Hoja 2 — Resumen por proveedor / origen.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compras"

    rojo        = PatternFill("solid", fgColor=ROJO)
    gris        = PatternFill("solid", fgColor=GRIS)
    ok_fill     = PatternFill("solid", fgColor=OK_COLOR)
    warn_fill   = PatternFill("solid", fgColor=WARN_COLOR)
    err_fill    = PatternFill("solid", fgColor=ERR_COLOR)
    azul_fill   = PatternFill("solid", fgColor=AZUL_CLARO)
    font_blanco = Font(bold=True, color="FFFFFF", size=12)
    font_bold   = Font(bold=True)
    font_titulo = Font(bold=True, color="FFFFFF", size=14)
    centro      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    izq         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    borde       = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    fecha = datetime.fromisoformat(
        datos.get("fecha", datetime.now().isoformat())
    ).strftime("%d/%m/%Y %H:%M:%S")

    # ── Título ────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:J1")
    ws["A1"] = "ARCAVISION — ARCA CONTINENTAL — REGISTRO DE COMPRAS"
    ws["A1"].font      = font_titulo
    ws["A1"].fill      = rojo
    ws["A1"].alignment = centro

    ws.row_dimensions[2].height = 22
    ws.merge_cells("A2:J2")
    ws["A2"] = (f"Proceso: {datos.get('objetivo','')}  |  "
                f"Origen: {datos.get('origen','')}  →  Destino: {datos.get('destino','')}  |  "
                f"Fecha: {fecha}")
    ws["A2"].fill      = PatternFill("solid", fgColor=GRIS)
    ws["A2"].alignment = centro

    ws.append([])  # fila 3 vacía

    # ── Headers ───────────────────────────────────────────────────────────────
    headers = [
        "N°", "Nombre del producto", "SKU / Código",
        "Cantidad", "Unidad", "Precio unitario",
        "Precio total", "Proveedor / Origen",
        "Sistema destino", "Estado",
    ]
    ws.row_dimensions[4].height = 24
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=4, column=col)
        c.font      = font_blanco
        c.fill      = rojo
        c.alignment = centro
        c.border    = borde

    # ── Filas de productos ────────────────────────────────────────────────────
    productos = _extraer_productos(datos)
    total_valor = 0.0

    for i, prod in enumerate(productos, 1):
        estado    = prod.get("estado", "ok")
        fill      = ok_fill if estado == "ok" else err_fill if estado == "error" else warn_fill
        qty       = prod.get("cantidad", 0)
        p_unit    = prod.get("precio_unitario", 0.0)
        p_total   = qty * p_unit if isinstance(qty, (int, float)) and isinstance(p_unit, (int, float)) else ""
        if isinstance(p_total, float):
            total_valor += p_total

        fila = [
            i,
            prod.get("nombre", "—"),
            prod.get("sku", "—"),
            qty,
            prod.get("unidad", "pza"),
            f"${p_unit:,.2f}" if isinstance(p_unit, float) else p_unit,
            f"${p_total:,.2f}" if isinstance(p_total, float) else "—",
            prod.get("proveedor", datos.get("origen", "—")),
            prod.get("sistema_destino", datos.get("destino", "—")),
            estado.upper(),
        ]
        row_num = ws.max_row + 1
        ws.row_dimensions[row_num].height = 20
        ws.append(fila)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row_num, column=col)
            c.fill      = fill
            c.border    = borde
            c.alignment = izq if col == 2 else centro

    # ── Fila de total ─────────────────────────────────────────────────────────
    ws.append([])
    total_row = ws.max_row + 1
    ws.row_dimensions[total_row].height = 22
    ws.append(["", "", "", "", "", "TOTAL COMPRA",
                f"${total_valor:,.2f}", "", "", ""])
    for col in range(1, 11):
        c = ws.cell(row=total_row, column=col)
        c.font = font_bold
        c.fill = PatternFill("solid", fgColor=GRIS)
        c.border = borde
        c.alignment = centro

    # ── Anchos de columna ─────────────────────────────────────────────────────
    anchos = {"A": 6, "B": 38, "C": 18, "D": 12, "E": 10,
              "F": 18, "G": 18, "H": 24, "I": 22, "J": 14}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w

    # ── Hoja 2: resumen por proveedor ─────────────────────────────────────────
    if productos:
        ws2 = wb.create_sheet("Resumen por proveedor")
        ws2.row_dimensions[1].height = 30
        ws2.merge_cells("A1:E1")
        ws2["A1"] = "RESUMEN POR PROVEEDOR / ORIGEN"
        ws2["A1"].font      = font_blanco
        ws2["A1"].fill      = rojo
        ws2["A1"].alignment = centro

        h2 = ["Proveedor", "N° productos", "Cantidad total", "Valor total ($)", "Estado predominante"]
        ws2.append(h2)
        for col in range(1, 6):
            c = ws2.cell(row=2, column=col)
            c.font = font_bold
            c.fill = PatternFill("solid", fgColor=GRIS)
            c.alignment = centro

        proveedores: dict = {}
        for prod in productos:
            prov = prod.get("proveedor", datos.get("origen", "—"))
            if prov not in proveedores:
                proveedores[prov] = {"n": 0, "qty": 0, "valor": 0.0, "estados": []}
            proveedores[prov]["n"]      += 1
            proveedores[prov]["qty"]    += prod.get("cantidad", 0) if isinstance(prod.get("cantidad"), int) else 0
            p_u = prod.get("precio_unitario", 0)
            q   = prod.get("cantidad", 0)
            if isinstance(p_u, (int, float)) and isinstance(q, (int, float)):
                proveedores[prov]["valor"] += p_u * q
            proveedores[prov]["estados"].append(prod.get("estado", "ok"))

        for prov, v in proveedores.items():
            estado_pred = Counter(v["estados"]).most_common(1)[0][0]
            fill = ok_fill if estado_pred == "ok" else err_fill if estado_pred == "error" else warn_fill
            row_n = ws2.max_row + 1
            ws2.append([prov, v["n"], v["qty"], f"${v['valor']:,.2f}", estado_pred.upper()])
            for col in range(1, 6):
                ws2.cell(row=row_n, column=col).fill = fill

        for col, w in zip(["A","B","C","D","E"], [30, 16, 16, 18, 22]):
            ws2.column_dimensions[col].width = w

    wb.save(output_path)

    # ── Cifrado del Excel financiero en reposo ────────────────────────────────
    # Si hay llave (ARCAVISION_ENC_KEY), el .xlsx con precios/totales se guarda
    # como .xlsx.enc (ilegible sin la llave) y se elimina el plano. Sin llave,
    # se deja el .xlsx normal (degradación elegante para demos).
    try:
        from postprocessing.crypto import cifrar_bytes, cifrado_disponible
        if cifrado_disponible():
            with open(output_path, "rb") as f:
                cifrado = cifrar_bytes(f.read())
            enc_path = output_path + ".enc"
            with open(enc_path, "wb") as f:
                f.write(cifrado)
            os.remove(output_path)
            print(f"  📦🔐 Excel compras cifrado: {enc_path}")
            return enc_path
    except Exception as e:
        print(f"  ⚠️  No se pudo cifrar el Excel ({e}) — se deja en claro")

    print(f"  📦 Excel compras: {output_path}")
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
# 2. PDF de reporte para ingenieros IA
# ═══════════════════════════════════════════════════════════════════════════════

def generar_pdf_reporte(datos: dict,
                        output_path: str = "reportes/reporte_ia_arcavision.pdf") -> str:
    """
    Genera un PDF técnico para el equipo de ingeniería con:
    - Resumen ejecutivo
    - Qué hizo la IA paso a paso
    - Errores encontrados y su contexto
    - Número de iteraciones / pasos
    - Métricas de rendimiento
    - Datos extraídos
    """
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
        title="ArcaVision — Reporte IA",
        author="ArcaVision · Arca Continental",
    )

    styles = getSampleStyleSheet()
    ROJO_RL_STR = "#C8102E"

    s_titulo = ParagraphStyle(
        "titulo", parent=styles["Heading1"],
        fontSize=20, textColor=colors.white,
        backColor=ROJO_RL, spaceAfter=0,
        spaceBefore=0, alignment=TA_CENTER,
        leftIndent=-20, rightIndent=-20,
        leading=28,
    )
    s_h2 = ParagraphStyle(
        "h2", parent=styles["Heading2"],
        fontSize=13, textColor=ROJO_RL,
        spaceBefore=14, spaceAfter=6,
        borderPad=4,
    )
    s_h3 = ParagraphStyle(
        "h3", parent=styles["Heading3"],
        fontSize=11, textColor=colors.HexColor("#4A4640"),
        spaceBefore=8, spaceAfter=4,
    )
    s_body = ParagraphStyle(
        "body", parent=styles["Normal"],
        fontSize=9.5, leading=14,
        textColor=colors.HexColor("#1C1A18"),
    )
    s_code = ParagraphStyle(
        "code", parent=styles["Code"],
        fontSize=8, leading=12,
        backColor=colors.HexColor("#F7F5F2"),
        leftIndent=10, rightIndent=10,
        borderPad=6,
    )
    s_meta = ParagraphStyle(
        "meta", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#8A857C"),
        alignment=TA_CENTER,
    )
    s_ok   = ParagraphStyle("ok",   parent=s_body, textColor=colors.HexColor("#1A7A45"))
    s_err  = ParagraphStyle("err",  parent=s_body, textColor=colors.HexColor("#C8102E"))
    s_warn = ParagraphStyle("warn", parent=s_body, textColor=colors.HexColor("#B87014"))

    fecha = datetime.fromisoformat(
        datos.get("fecha", datetime.now().isoformat())
    ).strftime("%d/%m/%Y %H:%M:%S")

    resultados  = datos.get("resultados", [])
    ok_steps    = [r for r in resultados if r.get("estado") == "ok"]
    err_steps   = [r for r in resultados if r.get("estado") == "error"]
    warn_steps  = [r for r in resultados if r.get("estado") == "advertencia"]
    total       = len(resultados)
    iteraciones = datos.get("iteraciones", total)

    story = []

    # ── Cabecera ──────────────────────────────────────────────────────────────
    story.append(Paragraph("⚡ ARCAVISION — ARCA CONTINENTAL", s_titulo))
    story.append(Paragraph("Reporte técnico de ejecución IA", s_titulo))
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph(
        f"Generado el {fecha}  ·  Motor: {datos.get('motor', 'browser_use / playwright')}  ·  "
        f"Versión: ArcaVision v12", s_meta,
    ))
    story.append(HRFlowable(width="100%", thickness=2, color=ROJO_RL, spaceAfter=8))

    # ── Resumen ejecutivo ─────────────────────────────────────────────────────
    story.append(Paragraph("1. Resumen ejecutivo", s_h2))
    resumen_data = [
        ["Proceso / objetivo", datos.get("objetivo", "—")],
        ["Sistema origen",     datos.get("origen",   "—")],
        ["Sistema destino",    datos.get("destino",  "—")],
        ["Fecha de ejecución", fecha],
        ["Total de pasos",     str(total)],
        ["Iteraciones del agente", str(iteraciones)],
        ["Pasos exitosos",     f"{len(ok_steps)} / {total}  ({_pct(len(ok_steps), total)}%)"],
        ["Advertencias",       f"{len(warn_steps)} / {total}  ({_pct(len(warn_steps), total)}%)"],
        ["Errores",            f"{len(err_steps)} / {total}  ({_pct(len(err_steps), total)}%)"],
        ["Motor de agente",    datos.get("motor", "browser_use")],
    ]
    t_res = Table(resumen_data, colWidths=[6*cm, 11*cm])
    t_res.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (0, -1), colors.HexColor("#F7F5F2")),
        ("FONTNAME",    (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1),
         [colors.HexColor("#F7F5F2"), colors.white]),
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(t_res)

    # ── Qué hizo la IA ────────────────────────────────────────────────────────
    story.append(Paragraph("2. Acciones ejecutadas por la IA", s_h2))
    story.append(Paragraph(
        f"El agente ejecutó <b>{total}</b> pasos en <b>{iteraciones}</b> iteraciones "
        f"para completar el proceso <i>{datos.get('objetivo','')}</i>. "
        f"A continuación el detalle cronológico:", s_body,
    ))
    story.append(Spacer(1, 0.3*cm))

    paso_headers = [["#", "Acción", "Intención / Descripción", "Estado", "Datos extraídos"]]
    paso_rows = []
    for r in resultados:
        est    = r.get("estado", "error")
        icono  = "✓" if est == "ok" else "✗" if est == "error" else "⚠"
        datos_ext = str(r.get("datos_extraidos", ""))[:80] if r.get("datos_extraidos") else "—"
        intencion = r.get("intencion", r.get("accion", "—"))[:60]
        paso_rows.append([
            str(r.get("paso", "")),
            r.get("accion", "—"),
            intencion,
            f"{icono} {est.upper()}",
            datos_ext,
        ])

    if paso_rows:
        t_pasos = Table(
            paso_headers + paso_rows,
            colWidths=[1*cm, 2.5*cm, 6.5*cm, 2.2*cm, 4.8*cm],
            repeatRows=1,
        )
        _estilo_tabla_pasos(t_pasos, paso_rows)
        story.append(t_pasos)

    # ── Errores y advertencias ────────────────────────────────────────────────
    story.append(Paragraph("3. Errores y advertencias encontrados", s_h2))

    problemas = err_steps + warn_steps
    if not problemas:
        story.append(Paragraph("✓ No se registraron errores ni advertencias en esta ejecución.", s_ok))
    else:
        story.append(Paragraph(
            f"Se detectaron <b>{len(err_steps)} error(es)</b> y "
            f"<b>{len(warn_steps)} advertencia(s)</b>.", s_body,
        ))
        story.append(Spacer(1, 0.2*cm))

        for r in problemas:
            est   = r.get("estado", "error")
            s_est = s_err if est == "error" else s_warn
            prefijo = "❌ ERROR" if est == "error" else "⚠️ ADVERTENCIA"
            story.append(KeepTogether([
                Paragraph(f"<b>{prefijo} — Paso {r.get('paso','?')} "
                          f"[{r.get('accion','?').upper()}]</b>", s_est),
                Paragraph(
                    f"Intención: {r.get('intencion', r.get('accion','—'))}",
                    s_body,
                ),
                Paragraph(
                    f"Detalle: {r.get('detalle_error', r.get('datos_extraidos', 'Sin detalle adicional'))}",
                    s_code,
                ),
                Spacer(1, 0.15*cm),
            ]))

    # ── Análisis de errores por tipo ──────────────────────────────────────────
    story.append(Paragraph("4. Clasificación de errores por tipo", s_h2))
    conteo = _clasificar_errores(resultados)
    if not conteo:
        story.append(Paragraph("Sin errores clasificados.", s_body))
    else:
        story.append(Paragraph(
            "Clasificación automática basada en el tipo de acción y estado de cada paso. "
            "Útil para identificar patrones de fallo recurrentes:", s_body,
        ))
        story.append(Spacer(1, 0.2*cm))
        t_err_headers = [["Tipo de error", "Descripción", "N° ocurrencias", "% del total"]]
        t_err_rows = [
            [tipo, desc, str(n), f"{_pct(n, max(total,1))}%"]
            for tipo, desc, n in conteo
        ]
        t_err = Table(
            t_err_headers + t_err_rows,
            colWidths=[4*cm, 7*cm, 3*cm, 3*cm],
            repeatRows=1,
        )
        t_err.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), ROJO_RL),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 9),
            ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("ROWBACKGROUNDS", (1, 0), (-1, -1),
             [colors.HexColor("#FDECEA"), colors.HexColor("#FFF5F6")]),
            ("ALIGN",         (2, 0), (-1, -1), "CENTER"),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ]))
        story.append(t_err)

    # ── Datos extraídos ───────────────────────────────────────────────────────
    datos_ext = datos.get("datos_extraidos", {})
    if datos_ext:
        story.append(Paragraph("5. Datos extraídos del portal", s_h2))
        story.append(Paragraph(
            "Información capturada automáticamente por el agente durante la ejecución:", s_body,
        ))
        story.append(Spacer(1, 0.2*cm))
        for clave, valor in datos_ext.items():
            story.append(Paragraph(f"<b>{clave}:</b>", s_h3))
            texto = str(valor)[:500] + ("…" if len(str(valor)) > 500 else "")
            story.append(Paragraph(texto, s_code))

    # ── Contexto técnico ──────────────────────────────────────────────────────
    story.append(Paragraph("6. Contexto técnico", s_h2))
    plan = datos.get("plan", {})
    story.append(Paragraph(
        f"<b>Motor:</b> {datos.get('motor', 'browser_use / playwright')}<br/>"
        f"<b>Modelo LLM:</b> claude-opus-4-5 (Anthropic)<br/>"
        f"<b>Pasos en el plan:</b> {len(plan.get('pasos', []))}<br/>"
        f"<b>Mapeo de campos:</b> {len(plan.get('mapeo_campos', []))} campos mapeados<br/>"
        f"<b>Credenciales utilizadas:</b> "
        f"{', '.join(plan.get('credenciales_obtenidas', {}).keys()) or 'Ninguna'}<br/>"
        f"<b>Excepciones declaradas en el plan:</b> {len(plan.get('excepciones', []))}",
        s_body,
    ))

    # ── Pie ───────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#DDDDDD")))
    story.append(Paragraph(
        "ArcaVision · Arca Continental · Hack4Her 2026 · Documento generado automáticamente",
        s_meta,
    ))

    doc.build(story)
    print(f"  📄 PDF reporte IA: {output_path}")
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
# 3. Excel de errores con gráficas
# ═══════════════════════════════════════════════════════════════════════════════

def generar_excel_errores(datos: dict,
                          output_path: str = "reportes/errores_arcavision.xlsx") -> str:
    """
    Genera un Excel con:
    - Hoja 1: tabla de errores por tipo + frecuencia
    - Hoja 2: tabla detallada de cada error
    - Gráfica de pastel y barras incrustada
    """
    wb    = openpyxl.Workbook()
    ws    = wb.active
    ws.title = "Errores por tipo"

    rojo        = PatternFill("solid", fgColor=ROJO)
    gris        = PatternFill("solid", fgColor=GRIS)
    err_fill    = PatternFill("solid", fgColor=ERR_COLOR)
    warn_fill   = PatternFill("solid", fgColor=WARN_COLOR)
    ok_fill     = PatternFill("solid", fgColor=OK_COLOR)
    font_blanco = Font(bold=True, color="FFFFFF", size=12)
    font_bold   = Font(bold=True)
    centro      = Alignment(horizontal="center", vertical="center")
    borde       = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    resultados = datos.get("resultados", [])
    fecha = datetime.fromisoformat(
        datos.get("fecha", datetime.now().isoformat())
    ).strftime("%d/%m/%Y %H:%M:%S")

    # ── Título ────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A1:F1")
    ws["A1"] = "ARCAVISION — ANÁLISIS DE ERRORES DEL AGENTE IA"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill      = rojo
    ws["A1"].alignment = centro

    ws.merge_cells("A2:F2")
    ws["A2"] = (f"Proceso: {datos.get('objetivo','')}  |  Fecha: {fecha}  |  "
                f"Total pasos: {len(resultados)}")
    ws["A2"].fill      = PatternFill("solid", fgColor=GRIS)
    ws["A2"].alignment = centro
    ws.append([])

    # ── Tabla de resumen por tipo ─────────────────────────────────────────────
    ws.append(["Tipo de error", "Descripción", "N° ocurrencias",
                "% del total", "Severidad", "Acción recomendada"])
    header_row = ws.max_row
    for col in range(1, 7):
        c = ws.cell(row=header_row, column=col)
        c.font      = font_blanco
        c.fill      = rojo
        c.alignment = centro
        c.border    = borde

    conteo = _clasificar_errores(resultados)
    total  = len(resultados)
    datos_para_grafica = []  # [(label, count)]

    for tipo, desc, n in conteo:
        pct      = _pct(n, max(total, 1))
        sev      = "ALTO" if n > total * 0.3 else "MEDIO" if n > total * 0.1 else "BAJO"
        accion   = _accion_recomendada(tipo)
        fill_row = err_fill if sev == "ALTO" else warn_fill if sev == "MEDIO" else ok_fill
        row_n    = ws.max_row + 1
        ws.row_dimensions[row_n].height = 18
        ws.append([tipo, desc, n, f"{pct}%", sev, accion])
        for col in range(1, 7):
            c = ws.cell(row=row_n, column=col)
            c.fill      = fill_row
            c.border    = borde
            c.alignment = centro if col != 2 else Alignment(wrap_text=True)
        datos_para_grafica.append((tipo, n))

    # ── Gráfica de barras ─────────────────────────────────────────────────────
    if datos_para_grafica:
        _insertar_graficas(ws, header_row + 1, len(conteo), len(resultados))

    # Anchos
    for col, w in zip(["A","B","C","D","E","F"], [22, 36, 16, 12, 12, 32]):
        ws.column_dimensions[col].width = w

    # ── Hoja 2: detalle de cada error ─────────────────────────────────────────
    ws2       = wb.create_sheet("Detalle de errores")
    ws2.row_dimensions[1].height = 30
    ws2.merge_cells("A1:G1")
    ws2["A1"] = "DETALLE CRONOLÓGICO DE ERRORES Y ADVERTENCIAS"
    ws2["A1"].font      = Font(bold=True, color="FFFFFF", size=12)
    ws2["A1"].fill      = rojo
    ws2["A1"].alignment = centro

    h3 = ["Paso", "Acción", "Estado", "Intención",
          "Detalle del error", "Tipo clasificado", "Timestamp"]
    ws2.append(h3)
    for col in range(1, len(h3) + 1):
        c = ws2.cell(row=2, column=col)
        c.font      = font_bold
        c.fill      = PatternFill("solid", fgColor=GRIS)
        c.alignment = centro

    ts = datos.get("fecha", datetime.now().isoformat())
    for r in resultados:
        est = r.get("estado", "ok")
        if est == "ok":
            continue
        fill_d = err_fill if est == "error" else warn_fill
        tipo_c = _tipo_error(r)
        row_n  = ws2.max_row + 1
        ws2.row_dimensions[row_n].height = 18
        ws2.append([
            r.get("paso", ""),
            r.get("accion", "—"),
            est.upper(),
            r.get("intencion", r.get("accion", "—"))[:60],
            str(r.get("detalle_error", r.get("datos_extraidos", "")))[:120],
            tipo_c,
            ts,
        ])
        for col in range(1, len(h3) + 1):
            ws2.cell(row=row_n, column=col).fill = fill_d

    for col, w in zip(["A","B","C","D","E","F","G"], [8, 16, 12, 38, 40, 22, 20]):
        ws2.column_dimensions[col].width = w

    # ── Hoja 3: resumen general ───────────────────────────────────────────────
    ws3       = wb.create_sheet("Resumen general")
    ws3.row_dimensions[1].height = 30
    ws3.merge_cells("A1:C1")
    ws3["A1"] = "RESUMEN GENERAL DE LA EJECUCIÓN"
    ws3["A1"].font      = Font(bold=True, color="FFFFFF", size=12)
    ws3["A1"].fill      = rojo
    ws3["A1"].alignment = centro

    n_ok   = sum(1 for r in resultados if r.get("estado") == "ok")
    n_err  = sum(1 for r in resultados if r.get("estado") == "error")
    n_warn = sum(1 for r in resultados if r.get("estado") == "advertencia")

    resumen_rows = [
        ["Métrica", "Valor", "Porcentaje"],
        ["Pasos totales",       total,  "100%"],
        ["Exitosos",            n_ok,   f"{_pct(n_ok,  max(total,1))}%"],
        ["Errores",             n_err,  f"{_pct(n_err, max(total,1))}%"],
        ["Advertencias",        n_warn, f"{_pct(n_warn,max(total,1))}%"],
        ["Tasa de éxito (%)",   f"{_pct(n_ok, max(total,1))}%", "—"],
    ]
    for i, row in enumerate(resumen_rows, 2):
        ws3.row_dimensions[i].height = 20
        ws3.append(row)
        fill_r = (rojo if i == 2 else
                  ok_fill if (len(row) > 0 and "Exitosos" in str(row[0])) else
                  err_fill if (len(row) > 0 and "Error" in str(row[0])) else
                  warn_fill if (len(row) > 0 and "Advert" in str(row[0])) else
                  PatternFill("solid", fgColor=GRIS))
        for col in range(1, 4):
            c = ws3.cell(row=i, column=col)
            c.fill = fill_r
            if i == 2:
                c.font = font_blanco
            c.alignment = centro

    # Gráfica de pastel en hoja 3
    if total > 0:
        pie = PieChart()
        pie.title = "Distribución de resultados"
        pie.style = 10
        pie.width  = 12
        pie.height = 10

        data_ref   = Reference(ws3, min_col=2, min_row=3, max_row=5)
        labels_ref = Reference(ws3, min_col=1, min_row=3, max_row=5)
        pie.add_data(data_ref)
        pie.set_categories(labels_ref)
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.fill import PatternFillProperties
        colores_pie = ["1A7A45", ROJO, "B87014"]
        pts = []
        for idx, hex_c in enumerate(colores_pie):
            dp = DataPoint(idx=idx)
            gp = GraphicalProperties()
            gp.solidFill = hex_c
            dp.spPr = gp
            pts.append(dp)
        pie.series[0].dPt = pts
        ws3.add_chart(pie, "E2")

    for col, w in zip(["A","B","C"], [24, 14, 14]):
        ws3.column_dimensions[col].width = w

    wb.save(output_path)
    print(f"  📊 Excel errores: {output_path}")
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _pct(n: int, total: int) -> int:
    return round(n / total * 100) if total else 0


def _extraer_productos(datos: dict) -> list:
    """
    Intenta extraer lista de productos de los datos_extraidos del agente.
    Si no hay estructura, construye una fila por resultado con datos disponibles.
    """
    productos = datos.get("productos", [])
    if productos:
        return productos

    # Intentar parsear datos_extraidos como JSON de productos
    for clave in ("datos_extraidos", "missing_products"):
        raw = datos.get(clave)
        if isinstance(raw, dict):
            for v in raw.values():
                try:
                    candidato = json.loads(v) if isinstance(v, str) else v
                    if isinstance(candidato, list) and candidato and "nombre" in candidato[0]:
                        return candidato
                except Exception:
                    pass
        if isinstance(raw, list):
            return raw

    # Fallback: una fila por paso con estado ok
    return [
        {
            "nombre":          r.get("intencion", r.get("accion", f"Producto paso {r.get('paso','')}"))[:60],
            "sku":             r.get("sku", "—"),
            "cantidad":        r.get("cantidad", 1),
            "unidad":          r.get("unidad", "pza"),
            "precio_unitario": r.get("precio_unitario", 0.0),
            "proveedor":       datos.get("origen", "—"),
            "sistema_destino": datos.get("destino", "—"),
            "estado":          r.get("estado", "ok"),
        }
        for r in datos.get("resultados", [])
        if r.get("estado") == "ok"
    ] or [{"nombre": "Sin productos registrados", "sku": "—", "cantidad": 0,
           "unidad": "—", "precio_unitario": 0.0, "estado": "advertencia"}]


def _tipo_error(r: dict) -> str:
    accion = r.get("accion", "").lower()
    det    = str(r.get("detalle_error", r.get("datos_extraidos", ""))).lower()
    if "timeout" in det:           return "Timeout"
    if "navegar" in accion:        return "Error de navegación"
    if "escribir" in accion:       return "Error al escribir"
    if "click" in accion:          return "Error de clic"
    if "seleccionar" in accion:    return "Error de selección"
    if "verificar" in accion:      return "Error de verificación"
    if "extraer" in accion:        return "Error de extracción"
    if r.get("estado") == "advertencia": return "Elemento no encontrado"
    return "Error desconocido"


def _clasificar_errores(resultados: list) -> list:
    """Retorna [(tipo, descripcion, count)] ordenado de mayor a menor."""
    descripciones = {
        "Timeout":                  "El agente esperó más tiempo del permitido",
        "Error de navegación":      "No se pudo cargar la URL o la página",
        "Error al escribir":        "No se pudo escribir en el campo indicado",
        "Error de clic":            "No se encontró el botón o elemento a clicar",
        "Error de selección":       "No se pudo seleccionar la opción indicada",
        "Error de verificación":    "La condición de verificación no se cumplió",
        "Error de extracción":      "No se pudo extraer el texto o dato esperado",
        "Elemento no encontrado":   "El agente no localizó el elemento en pantalla",
        "Error desconocido":        "Error no clasificado — revisar logs del agente",
    }
    conteo: Counter = Counter()
    for r in resultados:
        if r.get("estado") != "ok":
            conteo[_tipo_error(r)] += 1
    return [
        (tipo, descripciones.get(tipo, "—"), n)
        for tipo, n in conteo.most_common()
    ]


def _accion_recomendada(tipo: str) -> str:
    acciones = {
        "Timeout":               "Aumentar timeout o verificar velocidad de red",
        "Error de navegación":   "Verificar URL y disponibilidad del portal",
        "Error al escribir":     "Revisar selector visual del campo en el plan",
        "Error de clic":         "Ajustar descripción del botón en el plan",
        "Error de selección":    "Verificar opciones disponibles en el dropdown",
        "Error de verificación": "Revisar condición de éxito del paso",
        "Error de extracción":   "Verificar estructura de la página destino",
        "Elemento no encontrado":"Mejorar descripción visual del elemento",
        "Error desconocido":     "Revisar logs del agente para diagnóstico",
    }
    return acciones.get(tipo, "Revisar manualmente")


def _insertar_graficas(ws, data_start_row: int, n_rows: int, total: int):
    """Inserta gráfica de barras junto a la tabla de errores."""
    if n_rows < 1:
        return

    # Gráfica de barras
    bar = BarChart()
    bar.type    = "col"
    bar.style   = 10
    bar.title   = "Frecuencia de errores por tipo"
    bar.y_axis.title = "N° ocurrencias"
    bar.x_axis.title = "Tipo de error"
    bar.width   = 18
    bar.height  = 12

    data_ref   = Reference(ws, min_col=3, min_row=data_start_row,
                           max_row=data_start_row + n_rows - 1)
    cats_ref   = Reference(ws, min_col=1, min_row=data_start_row,
                           max_row=data_start_row + n_rows - 1)
    bar.add_data(data_ref, titles_from_data=False)
    bar.set_categories(cats_ref)
    bar.series[0].graphicalProperties.solidFill = ROJO

    anchor_row = data_start_row + n_rows + 3
    ws.add_chart(bar, f"A{anchor_row}")




# ═══════════════════════════════════════════════════════════════════════════════
# Función unificada — genera los tres artefactos
# ═══════════════════════════════════════════════════════════════════════════════

def generar_todos_los_reportes(datos: dict) -> dict:
    """
    Punto de entrada principal. Genera los tres artefactos y retorna sus rutas.
    """
    rutas = {}
    try:
        rutas["excel_compras"] = generar_excel_compras(datos)
    except Exception as e:
        print(f"  ⚠️  Excel compras: {e}")
        rutas["excel_compras"] = None

    try:
        rutas["pdf_reporte"] = generar_pdf_reporte(datos)
    except Exception as e:
        print(f"  ⚠️  PDF reporte: {e}")
        rutas["pdf_reporte"] = None

    try:
        rutas["excel_errores"] = generar_excel_errores(datos)
    except Exception as e:
        print(f"  ⚠️  Excel errores: {e}")
        rutas["excel_errores"] = None

    return rutas


def _estilo_tabla_pasos(tabla, rows):
    n = len(rows)
    estilo = [
        ("BACKGROUND",  (0, 0), (-1, 0), ROJO_RL),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 8),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]
    for i, r in enumerate(rows, 1):
        est = r[3] if len(r) > 3 else ""
        if "ERROR" in est or "✗" in est:
            estilo.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FDECEA")))
        elif "ADVERT" in est or "⚠" in est:
            estilo.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FEF3E2")))
        else:
            bg = colors.HexColor("#F7FAF8") if i % 2 == 0 else colors.white
            estilo.append(("BACKGROUND", (0, i), (-1, i), bg))
    tabla.setStyle(TableStyle(estilo))


# ═══════════════════════════════════════════════════════════════════════════════
# Mantener compatibilidad con código existente
# ═══════════════════════════════════════════════════════════════════════════════

def generar_excel(datos: dict, output_path: str = "reportes/reporte_arcavision.xlsx") -> str:
    """Alias de compatibilidad — genera el Excel de compras."""
    return generar_excel_compras(datos, output_path)


def generar_ticket_html(datos: dict) -> str:
    fecha    = datetime.fromisoformat(datos.get("fecha", datetime.now().isoformat())).strftime("%d/%m/%Y %H:%M")
    objetivo = datos.get("objetivo", "Proceso automatizado")
    origen   = datos.get("origen", "")
    destino  = datos.get("destino", "")
    resultados = datos.get("resultados", [])
    ok    = sum(1 for r in resultados if r.get("estado") == "ok")
    total = len(resultados)

    filas_pasos = ""
    for r in resultados:
        estado = r.get("estado", "error")
        color  = "#E6F4EC" if estado == "ok" else "#FDECEA" if estado == "error" else "#FEF3E2"
        icono  = "✅" if estado == "ok" else "❌" if estado == "error" else "⚠️"
        datos_ext = str(r.get("datos_extraidos", ""))[:100] if r.get("datos_extraidos") else ""
        filas_pasos += f"""
        <tr style="background:{color}">
          <td style="padding:8px;border-bottom:1px solid #eee;text-align:center">{r.get('paso','')}</td>
          <td style="padding:8px;border-bottom:1px solid #eee">{r.get('accion','')}</td>
          <td style="padding:8px;border-bottom:1px solid #eee;text-align:center">{icono} {estado.upper()}</td>
          <td style="padding:8px;border-bottom:1px solid #eee;font-size:12px;color:#666">{datos_ext}</td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>ArcaVision — Ticket</title></head>
<body style="margin:0;background:#f5f5f5;font-family:'Segoe UI',Arial,sans-serif">
<div style="max-width:640px;margin:30px auto;background:white;border-radius:8px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.1)">
  <div style="background:#C8102E;padding:28px;text-align:center">
    <h1 style="color:white;margin:0;font-size:22px;letter-spacing:1px">⚡ ArcaVision</h1>
    <p style="color:#ffcccc;margin:6px 0 0;font-size:14px">Arca Continental · Reporte de ejecución</p>
  </div>
  <div style="padding:28px">
    <table style="width:100%;margin-bottom:20px">
      <tr><td style="color:#888;font-size:12px;text-transform:uppercase">Proceso</td></tr>
      <tr><td style="font-size:16px;font-weight:600">{objetivo}</td></tr>
    </table>
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-bottom:24px">
      <div style="background:#f9f9f9;border-radius:6px;padding:14px;text-align:center">
        <div style="font-size:11px;color:#888;text-transform:uppercase;margin-bottom:4px">Fecha</div>
        <div style="font-size:13px;font-weight:600">{fecha}</div>
      </div>
      <div style="background:#f9f9f9;border-radius:6px;padding:14px;text-align:center">
        <div style="font-size:11px;color:#888;text-transform:uppercase;margin-bottom:4px">Resultado</div>
        <div style="font-size:20px;font-weight:700;color:{'#1A7A45' if ok==total else '#C8102E'}">{ok}/{total}</div>
      </div>
      <div style="background:#f9f9f9;border-radius:6px;padding:14px;text-align:center">
        <div style="font-size:11px;color:#888;text-transform:uppercase;margin-bottom:4px">Flujo</div>
        <div style="font-size:12px;font-weight:600">{origen} → {destino}</div>
      </div>
    </div>
    <table style="width:100%;border-collapse:collapse">
      <thead><tr style="background:#f2f2f2">
        <th style="padding:8px;text-align:center;font-size:12px">#</th>
        <th style="padding:8px;text-align:left;font-size:12px">Acción</th>
        <th style="padding:8px;text-align:center;font-size:12px">Estado</th>
        <th style="padding:8px;text-align:left;font-size:12px">Datos</th>
      </tr></thead>
      <tbody>{filas_pasos}</tbody>
    </table>
  </div>
  <div style="background:#f2f2f2;padding:14px;text-align:center;font-size:11px;color:#999">
    ArcaVision · Arca Continental · Hack4Her 2026
  </div>
</div>
</body></html>"""


def guardar_ticket(html: str, output_path: str = "reportes/ticket.html") -> str:
    Path("reportes").mkdir(exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    return output_path


HISTORIAL_PATH = "reportes/historial_arca.xlsx"

def agregar_al_historial_excel(datos: dict):
    Path("reportes").mkdir(exist_ok=True)
    rojo        = PatternFill("solid", fgColor=ROJO)
    font_blanco = Font(bold=True, color="FFFFFF", size=11)
    centro      = Alignment(horizontal="center")

    if Path(HISTORIAL_PATH).exists():
        wb = openpyxl.load_workbook(HISTORIAL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial Pedidos"
        headers = ["Fecha","Objetivo","Origen","Destino","Pasos","Exitosos","Errores","Motor"]
        ws.append(headers)
        for col in range(1, len(headers)+1):
            c = ws.cell(row=1, column=col)
            c.font = font_blanco; c.fill = rojo; c.alignment = centro
        for ltr in ["A","B","C","D"]:
            ws.column_dimensions[ltr].width = 24

    resultados = datos.get("resultados", [])
    ok  = sum(1 for r in resultados if r.get("estado") == "ok")
    err = sum(1 for r in resultados if r.get("estado") == "error")
    fecha = datetime.fromisoformat(
        datos.get("fecha", datetime.now().isoformat())
    ).strftime("%d/%m/%Y %H:%M")
    ws.append([fecha, datos.get("objetivo",""), datos.get("origen",""),
               datos.get("destino",""), len(resultados), ok, err,
               datos.get("motor","playwright")])
    wb.save(HISTORIAL_PATH)
    return HISTORIAL_PATH


# ═══════════════════════════════════════════════════════════════════════════════
# SISTEMA DE TICKETS — extracción, email y persistencia
# (integrado desde ArcaVision-Izhar con mejoras para el contexto FastAPI)
# ═══════════════════════════════════════════════════════════════════════════════

def extraer_datos_pedido(resultado_texto: str, datos_reporte: dict) -> dict:
    """
    Extrae productos, totales y datos de envío del texto libre del agente.
    Usa regex del ZIP original + fallback a datos estructurados del agente.
    """
    # ── 1. Productos ya estructurados por el agente (camino feliz) ────────────
    productos = datos_reporte.get("productos") or []

    # ── 2. Fallback regex sobre el texto del agente ───────────────────────────
    if not productos and resultado_texto:
        nombres_vistos: set = set()
        patron = r"\d+\.\s+(.+?)\s+-\s+\$([0-9,]+\.?[0-9]*)"
        for m in re.finditer(patron, resultado_texto):
            nombre = m.group(1).strip()
            precio = float(m.group(2).replace(",", ""))
            if nombre not in nombres_vistos:
                nombres_vistos.add(nombre)
                productos.append({"nombre": nombre, "precio_unitario": precio,
                                   "cantidad": 1, "sku": "", "estado": "ok"})

    # ── 3. Totales con regex ──────────────────────────────────────────────────
    def _monto(patron_re: str) -> float:
        m = re.search(patron_re, resultado_texto or "")
        return float(m.group(1).replace(",", "")) if m else 0.0

    subtotal  = _monto(r"[Ss]ubtotal.*?\$([0-9,]+\.?[0-9]*)")
    impuestos = _monto(r"[Ii]mpuesto.*?\$([0-9,]+\.?[0-9]*)")
    total     = _monto(r"TOTAL.*?\$([0-9,]+\.?[0-9]*)")
    if total == 0 and productos:
        total = sum(p.get("precio_unitario", 0) * p.get("cantidad", 1) for p in productos)

    # ── 4. Datos de envío ─────────────────────────────────────────────────────
    cliente_m = re.search(r"[Dd]irecci[oó]n:\s+(.+?),", resultado_texto or "")
    zip_m     = re.search(r"[Cc][óo]digo\s+[Pp]ostal\s+(\d+)", resultado_texto or "")
    envio_m   = re.search(r"[Mm][eé]todo:\s+(.+?)(?:\n|$)", resultado_texto or "")

    return {
        "productos":  productos,
        "subtotal":   subtotal,
        "impuestos":  impuestos,
        "total":      total,
        "cliente":    (cliente_m.group(1).strip() if cliente_m
                       else datos_reporte.get("email", "Cliente")),
        "zip":        zip_m.group(1) if zip_m else "",
        "envio":      (envio_m.group(1).strip() if envio_m else "Envío estándar"),
        "fecha":      datos_reporte.get("fecha", datetime.now().isoformat()),
        "objetivo":   datos_reporte.get("objetivo", "Pedido"),
        "comercio":   datos_reporte.get("origen", ""),
    }


def generar_ticket_html_premium(datos_pedido: dict, datos_reporte: dict) -> str:
    """
    Ticket HTML de alta calidad — diseño Arca Continental.
    Muestra productos reales, totales, pasos del proceso y estado final.
    """
    fecha    = datetime.fromisoformat(
        datos_pedido.get("fecha", datetime.now().isoformat())
    ).strftime("%d/%m/%Y %H:%M")
    cliente  = datos_pedido.get("cliente", "Cliente")
    objetivo = datos_pedido.get("objetivo", "Pedido automatizado")
    envio    = datos_pedido.get("envio", "—")
    zip_code = datos_pedido.get("zip", "")

    resultados = datos_reporte.get("resultados", [])
    ok_count   = sum(1 for r in resultados if r.get("estado") == "ok")
    total_pasos = len(resultados)

    # Filas de productos
    productos = datos_pedido.get("productos", [])
    filas_prod = ""
    if productos:
        for p in productos:
            precio  = p.get("precio_unitario", p.get("precio", 0))
            cant    = p.get("cantidad", 1)
            subtot  = precio * cant
            filas_prod += f"""
            <tr>
              <td style="padding:10px 12px;border-bottom:1px solid #F2ECE8">{p.get('nombre','—')}</td>
              <td style="padding:10px 12px;border-bottom:1px solid #F2ECE8;text-align:center">{cant}</td>
              <td style="padding:10px 12px;border-bottom:1px solid #F2ECE8;text-align:right">${precio:,.2f}</td>
              <td style="padding:10px 12px;border-bottom:1px solid #F2ECE8;text-align:right;font-weight:600">${subtot:,.2f}</td>
            </tr>"""
    else:
        filas_prod = "<tr><td colspan='4' style='padding:14px;text-align:center;color:#999'>Sin productos detallados</td></tr>"

    # Filas de pasos (resumen compacto)
    filas_pasos = ""
    for r in resultados[:10]:
        estado = r.get("estado", "error")
        icono  = "✅" if estado == "ok" else "❌" if estado == "error" else "⚠️"
        color  = "#E6F4EC" if estado == "ok" else "#FDECEA" if estado == "error" else "#FEF3E2"
        filas_pasos += f"""
        <tr style="background:{color}">
          <td style="padding:6px 10px;font-size:12px;color:#666">{r.get('paso','')}</td>
          <td style="padding:6px 10px;font-size:12px">{r.get('intencion',r.get('accion',''))[:55]}</td>
          <td style="padding:6px 10px;font-size:12px;text-align:center">{icono}</td>
        </tr>"""

    estado_color = "#1A7A45" if ok_count == total_pasos else "#C8102E"
    estado_texto = "COMPLETADO" if ok_count == total_pasos else f"{ok_count}/{total_pasos} pasos"

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>ArcaVision — Ticket de pedido</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
</head>
<body style="margin:0;background:#F5F0EB;font-family:'Segoe UI',Arial,sans-serif;padding:20px">
<div style="max-width:620px;margin:0 auto;background:white;border-radius:12px;overflow:hidden;
            box-shadow:0 4px 24px rgba(28,26,24,.14)">

  <!-- Header -->
  <div style="background:linear-gradient(135deg,#C8102E 0%,#a00e24 100%);padding:32px;text-align:center;position:relative">
    <div style="font-size:11px;letter-spacing:3px;color:rgba(255,255,255,.6);text-transform:uppercase;margin-bottom:6px">Arca Continental</div>
    <h1 style="color:white;margin:0;font-size:26px;letter-spacing:2px;font-weight:800">⚡ ARCAVISION</h1>
    <p style="color:rgba(255,255,255,.7);margin:8px 0 0;font-size:13px">Confirmación de pedido automatizado</p>
    <div style="margin-top:18px;display:inline-block;padding:6px 20px;border-radius:20px;
                background:rgba(255,255,255,.15);color:white;font-size:12px;letter-spacing:1px">
      {fecha}
    </div>
  </div>

  <!-- Estado banner -->
  <div style="background:{estado_color};padding:12px 28px;display:flex;align-items:center;justify-content:space-between">
    <span style="color:white;font-size:13px;font-weight:600">{objetivo}</span>
    <span style="color:white;font-size:13px;font-weight:700;letter-spacing:1px">{estado_texto}</span>
  </div>

  <div style="padding:28px">

    <!-- Info cliente y envío -->
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:24px">
      <div style="background:#FBF8F5;border-radius:8px;padding:16px;border-left:3px solid #C8102E">
        <div style="font-size:10px;color:#999;text-transform:uppercase;letter-spacing:1px;margin-bottom:4px">Cliente</div>
        <div style="font-weight:700;color:#1C1A18">{cliente}</div>
      </div>
      <div style="background:#FBF8F5;border-radius:8px;padding:16px;border-left:3px solid #C9A35B">
        <div style="font-size:10px;color:#999;text-transform:uppercase;letter-spacing:1px;margin-bottom:4px">Envío · CP {zip_code}</div>
        <div style="font-weight:700;color:#1C1A18">{envio}</div>
      </div>
    </div>

    <!-- Tabla productos -->
    <div style="margin-bottom:24px">
      <div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:1px;
                  color:#999;margin-bottom:10px;padding-bottom:8px;border-bottom:2px solid #F2ECE8">
        Productos del pedido
      </div>
      <table style="width:100%;border-collapse:collapse">
        <thead>
          <tr style="background:#FBF8F5">
            <th style="padding:10px 12px;text-align:left;font-size:11px;color:#888">Producto</th>
            <th style="padding:10px 12px;text-align:center;font-size:11px;color:#888">Cant.</th>
            <th style="padding:10px 12px;text-align:right;font-size:11px;color:#888">Precio</th>
            <th style="padding:10px 12px;text-align:right;font-size:11px;color:#888">Total</th>
          </tr>
        </thead>
        <tbody>{filas_prod}</tbody>
      </table>
      <!-- Totales -->
      <div style="margin-top:12px;border-top:2px solid #F2ECE8;padding-top:12px">
        <div style="display:flex;justify-content:space-between;padding:4px 12px;font-size:13px;color:#666">
          <span>Subtotal</span><span>${datos_pedido.get('subtotal',0):,.2f}</span>
        </div>
        <div style="display:flex;justify-content:space-between;padding:4px 12px;font-size:13px;color:#666">
          <span>Impuestos</span><span>${datos_pedido.get('impuestos',0):,.2f}</span>
        </div>
        <div style="display:flex;justify-content:space-between;padding:10px 12px;
                    font-size:17px;font-weight:800;color:#C8102E;
                    border-top:2px solid #F2ECE8;margin-top:6px">
          <span>TOTAL</span><span>${datos_pedido.get('total',0):,.2f}</span>
        </div>
      </div>
    </div>

    <!-- Resumen de pasos (si hay) -->
    {"" if not filas_pasos else f'''
    <div style="margin-bottom:24px">
      <div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:1px;
                  color:#999;margin-bottom:10px;padding-bottom:8px;border-bottom:2px solid #F2ECE8">
        Ejecución del agente
      </div>
      <table style="width:100%;border-collapse:collapse;font-size:12px">
        <thead><tr style="background:#FBF8F5">
          <th style="padding:6px 10px;text-align:left;font-size:10px;color:#999">#</th>
          <th style="padding:6px 10px;text-align:left;font-size:10px;color:#999">Paso</th>
          <th style="padding:6px 10px;text-align:center;font-size:10px;color:#999">Estado</th>
        </tr></thead>
        <tbody>{filas_pasos}</tbody>
      </table>
    </div>'''}

  </div>

  <!-- Footer -->
  <div style="background:#1C1A18;padding:16px 28px;display:flex;align-items:center;justify-content:space-between">
    <div style="font-size:11px;color:rgba(255,255,255,.4);letter-spacing:1px">
      ArcaVision · Arca Continental
    </div>
    <div style="font-size:11px;color:rgba(255,255,255,.3)">Hack4Her 2026</div>
  </div>
</div>
</body></html>"""


def enviar_ticket(datos_pedido: dict, datos_reporte: dict,
                  email_cliente: str, excel_path: str = None) -> str:
    """
    Envía el ticket premium por email con el Excel adjunto.
    Siempre guarda el HTML localmente aunque el email falle.

    Requiere en .env:
      EMAIL_REMITENTE=tu_cuenta@gmail.com
      EMAIL_PASSWORD=xxxx xxxx xxxx xxxx   ← App Password de Gmail (16 chars)

    El ticket se envía a `email_cliente` (el email del usuario que inicia sesión).
    """
    remitente = os.getenv("EMAIL_REMITENTE", "").strip()
    password  = os.getenv("EMAIL_PASSWORD",  "").strip()

    html        = generar_ticket_html_premium(datos_pedido, datos_reporte)
    ticket_path = "reportes/ticket.html"
    Path("reportes").mkdir(exist_ok=True)
    with open(ticket_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  🎫 Ticket HTML guardado: {ticket_path}")

    if not remitente or not password:
        print("  ℹ️  EMAIL_REMITENTE / EMAIL_PASSWORD no configurados — solo guardado local")
        datos_pedido["ticket_enviado"] = False
        return ticket_path
    if not email_cliente:
        print("  ℹ️  Sin email de cliente — solo guardado local")
        datos_pedido["ticket_enviado"] = False
        return ticket_path

    fecha  = datetime.fromisoformat(
        datos_pedido.get("fecha", datetime.now().isoformat())
    ).strftime("%d/%m/%Y %H:%M")
    total  = datos_pedido.get("total", 0)
    n_prod = len(datos_pedido.get("productos", []))

    # Estructura correcta: multipart/mixed con alternative interior para HTML
    msg = MIMEMultipart("mixed")
    msg["From"]    = remitente
    msg["To"]      = email_cliente
    msg["Subject"] = (f"✅ Pedido Arca Continental — {fecha} — "
                      f"{n_prod} producto(s) — ${total:,.2f}")

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(html, "html", "utf-8"))
    msg.attach(alt)

    if excel_path and Path(excel_path).exists():
        fecha_safe = fecha[:10].replace("/", "-")
        with open(excel_path, "rb") as f:
            contenido = f.read()

        # El Excel en reposo ya viene cifrado (.xlsx.enc). Para adjuntarlo legible
        # al cliente lo desciframos en memoria; el correo viaja cifrado por TLS.
        if str(excel_path).endswith(".enc"):
            try:
                from postprocessing.crypto import descifrar_bytes
                contenido = descifrar_bytes(contenido)
            except Exception as e:
                print(f"  ⚠️  No se pudo descifrar el Excel para adjuntar ({e})")

        # Si EMAIL_CIFRAR_ADJUNTO=1 y hay llave, se re-cifra el adjunto del correo
        # (se envía .xlsx.enc, ilegible sin la llave) en vez del Excel legible.
        cifrar_adj = os.getenv("EMAIL_CIFRAR_ADJUNTO", "").strip() in ("1", "true", "yes")
        if cifrar_adj:
            try:
                from postprocessing.crypto import cifrar_bytes, cifrado_disponible
                if cifrado_disponible():
                    contenido = cifrar_bytes(contenido)
                    filename = f"pedido_arca_{fecha_safe}.xlsx.enc"
                else:
                    print("  ⚠️  EMAIL_CIFRAR_ADJUNTO activo pero sin llave — se adjunta sin cifrar")
                    filename = f"pedido_arca_{fecha_safe}.xlsx"
            except Exception as e:
                print(f"  ⚠️  No se pudo cifrar el adjunto ({e}) — se adjunta sin cifrar")
                filename = f"pedido_arca_{fecha_safe}.xlsx"
        else:
            filename = f"pedido_arca_{fecha_safe}.xlsx"

        part = MIMEBase("application", "octet-stream")
        part.set_payload(contenido)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={filename}")
        msg.attach(part)

    enviado = False
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=30) as server:
            server.login(remitente, password)
            server.sendmail(remitente, email_cliente, msg.as_string())
        print(f"  📧 Ticket enviado a {email_cliente}")
        enviado = True
    except smtplib.SMTPAuthenticationError:
        print("  ❌ Email: autenticación fallida — verifica EMAIL_PASSWORD (usa App Password de Gmail, no tu contraseña normal)")
    except smtplib.SMTPException as e:
        print(f"  ❌ Email SMTP: {e}")
    except Exception as e:
        print(f"  ⚠️  Email no enviado: {e}")

    datos_pedido["ticket_enviado"] = enviado
    return ticket_path


def agregar_a_google_sheets(datos_pedido: dict):
    """
    Añade la fila del pedido a Google Sheets (opcional — requiere credenciales).
    Falla silenciosamente si no hay google_credentials.json.
    """
    creds_path = os.getenv("GOOGLE_CREDENTIALS_PATH", "google_credentials.json")
    sheet_id   = os.getenv("GOOGLE_SHEET_ID", "")
    if not Path(creds_path).exists() or not sheet_id:
        return False
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        scopes  = ["https://www.googleapis.com/auth/spreadsheets",
                   "https://www.googleapis.com/auth/drive"]
        creds   = Credentials.from_service_account_file(creds_path, scopes=scopes)
        client  = gspread.authorize(creds)
        sheet   = client.open_by_key(sheet_id).sheet1
        if sheet.row_count == 0 or sheet.cell(1, 1).value != "Fecha":
            sheet.append_row(["Fecha","Comercio","Cliente","CP","Productos",
                               "Subtotal","Impuestos","Total","Envío"])
        fecha         = datetime.fromisoformat(
            datos_pedido.get("fecha", datetime.now().isoformat())
        ).strftime("%d/%m/%Y %H:%M")
        productos_txt = ", ".join([
            f"{p.get('nombre','')} (${p.get('precio_unitario',p.get('precio',0)):.2f})"
            for p in datos_pedido.get("productos", [])
        ])
        sheet.append_row([
            fecha,
            datos_pedido.get("comercio", datos_pedido.get("objetivo", "")),
            datos_pedido.get("cliente", ""),
            datos_pedido.get("zip", ""),
            productos_txt,
            datos_pedido.get("subtotal", 0),
            datos_pedido.get("impuestos", 0),
            datos_pedido.get("total", 0),
            datos_pedido.get("envio", ""),
        ])
        print("  📊 Google Sheets actualizado")
        return True
    except Exception as e:
        print(f"  ℹ️  Google Sheets no disponible: {e}")
        return False


def procesar_ticket_completo(datos_reporte: dict,
                             email_cliente: str,
                             excel_path: str = None,
                             sesion_id: int = None) -> dict:
    """
    Pipeline completo de ticket post-ejecución:
      1. Extrae datos de pedido del texto + estructura del agente
      2. Genera y envía ticket HTML por email con Excel adjunto
      3. Guarda en SQLite tabla pedidos (con el estado real del envío)
      4. Actualiza historial Excel acumulativo
      5. Sube a Google Sheets (si está configurado)
    """
    # ── Texto libre del agente para extracción regex ──────────────────────────
    # Incluye el resumen final del agente + todos los datos_extraidos por paso
    partes = []
    for r in datos_reporte.get("resultados", []):
        ext = r.get("datos_extraidos")
        if isinstance(ext, dict):
            partes.append(str(ext.get("resumen", "")) + " " + str(ext))
        elif ext:
            partes.append(str(ext))
    resultado_texto = " ".join(partes)

    datos_pedido = extraer_datos_pedido(resultado_texto, datos_reporte)

    # ── Anclaje de integridad en Solana (hash del registro financiero) ────────
    # Complementa al cifrado: el cifrado oculta el dato, el hash on-chain prueba
    # que no fue alterado. Usa fallback local si Solana no está configurado.
    try:
        from postprocessing.solana_audit import registrar_en_solana_safe
        registro_financiero = {
            "cliente":   datos_pedido.get("cliente", ""),
            "comercio":  datos_pedido.get("comercio", ""),
            "productos": datos_pedido.get("productos", []),
            "subtotal":  datos_pedido.get("subtotal", 0),
            "impuestos": datos_pedido.get("impuestos", 0),
            "total":     datos_pedido.get("total", 0),
            "fecha":     datos_pedido.get("fecha", ""),
        }
        anclaje = registrar_en_solana_safe(registro_financiero, fallback_on_error=True)
        datos_pedido["solana_hash"]    = anclaje.get("payload_hash")
        datos_pedido["solana_tx"]      = anclaje.get("tx_signature")
        datos_pedido["solana_explorer"] = anclaje.get("explorer_url")
    except Exception as e:
        print(f"  ⚠️  Anclaje Solana no disponible: {e}")

    # ── Ticket email (primero: define datos_pedido["ticket_enviado"]) ─────────
    ticket_path = enviar_ticket(datos_pedido, datos_reporte, email_cliente, excel_path)

    # ── SQLite (ya con el estado real del envío) ──────────────────────────────
    try:
        from database.db import guardar_pedido
        pedido_id = guardar_pedido(datos_pedido, sesion_id=sesion_id)
        print(f"  💾 Pedido #{pedido_id} guardado en SQLite")
    except Exception as e:
        print(f"  ⚠️  SQLite pedido: {e}")
        pedido_id = None

    # ── Historial Excel acumulativo ───────────────────────────────────────────
    try:
        agregar_al_historial_excel(datos_reporte)
    except Exception as e:
        print(f"  ⚠️  Historial Excel: {e}")

    # ── Google Sheets ─────────────────────────────────────────────────────────
    agregar_a_google_sheets(datos_pedido)

    return {"datos_pedido": datos_pedido, "ticket_path": ticket_path, "pedido_id": pedido_id}
